import json
import uuid

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Avg
from django.shortcuts import redirect
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from collections import defaultdict
from io import BytesIO

from basic.models import User


def doctor_login_view(request):
    return render(request, 'basic/dashboard/doctor_login.html')



def doctor_user_login(request):
    if request.method == "POST":
        # Getting the username and password in the form post in the html page/
        username = request.POST.get("username")
        password = request.POST.get("password")

        # Authenticate the username and password
        user = authenticate(request, username=username, password=password)

        if user is not None:
            # Authentication is successful
            login(request, user)
            return redirect("basic:doctor-dashboard")  # Redirecting to the dashboard
        else:
            # Authentication is denied
            messages.error(request, "Invalid username or password")
            return  redirect("basic:doctor-login") # Redirecting to the login page, when refreshed

    return render(request, "basic/dashboard/doctor_login.html")

@login_required(login_url='/basic/')
def doctor_dashboard(request):
    test_sessions = TestSession.objects.filter(doctor=request.user)

    completed_tests = test_sessions.count()
    active_tests = test_sessions.filter(duration__isnull=False).count()
    pending_tests = test_sessions.filter(duration__isnull=True).count()

    # Get recent 10 completed tests
    recent_tests = test_sessions.filter(duration__isnull=False).order_by('-date')[:10]

    # Chart data: language distribution
    language_counts = test_sessions.values('language').annotate(count=Count('test_id'))

    chart_labels = [entry['language'] for entry in language_counts]
    chart_data = [entry['count'] for entry in language_counts]

    return render(request, 'basic/dashboard/doctor_dashboard.html', {
        'completed_tests': completed_tests,
        'active_tests': active_tests,
        'pending_tests': pending_tests,
        'recent_tests': recent_tests,
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
    })


@login_required(login_url='/basic/')
def doctor_results(request):
    test_sessions = TestSession.objects.filter(doctor=request.user).order_by('-date')
    return render(request, 'basic/dashboard/doctor_results.html', {'test_sessions': test_sessions})

@login_required(login_url='/basic/')
def doctor_statistics(request):
    return render(request, 'basic/dashboard/doctor_statistics.html')

@login_required(login_url='/basic/')
def doctor_newtest(request):
    return render(request, "basic/dashboard/doctor_newtest.html")

def base(request):
    return render(request, "basic/dashboard/base.html")

def landing(request):
     total_tests = TestSession.objects.count()  # Count all test sessions
     total_doctors = User.objects.filter(testsession__isnull=False).distinct().count()  # Count unique doctors with tests

     return render(request, "basic/dashboard/landing.html", {
        "total_tests": total_tests,
        "total_doctors": total_doctors,
    })
def doctor_logout_view(request):
    logout(request)
    return redirect('basic:landing')

import random
from .models import TestSession, Stimuli


def generate_test(request):
    age = request.GET.get("age")
    language = request.GET.get("language")
    if age is None:
        return JsonResponse({"error": "Age parameter is required."}, status=400)

    # Define preset test orders (4 possible sequences)
    test_orders = [
        [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11],
        [3, 2, 1, 0, 7, 6, 5, 4, 11, 10, 9, 8],
        [6, 7, 8, 9, 10, 11, 0, 1, 2, 3, 4, 5],
        [11, 10, 9, 8, 7, 6, 5, 4, 3, 2, 1, 0]
    ]

    # Pick a random test order
    selected_order = random.choice(test_orders)

    # Fetch stimuli and apply the selected order
    stimuli_list = list(Stimuli.objects.exclude(type="Practice"))

    ordered_stimuli = [stimuli_list[i] for i in selected_order]

    # Store the order as a comma-separated string
    stimuli_order_str = ",".join(map(str, selected_order))

    test_uuid = str(uuid.uuid4())

    # Create the test session
    test_session = TestSession.objects.create(
        doctor=request.user,
        age=age,
        language=language,
        stimuli_order=stimuli_order_str,
        test_id=test_uuid,
        date=timezone.now(),
    )

    # Generate responses
    responses = []
    for stimulus in ordered_stimuli:
        response = Response.objects.create(
            test=test_session,
            stim=stimulus,
        )
        responses.append({
            "response_id": response.response_id,
            "stimulus": stimulus.stim_id,
        })

    if language == 'en':
        test_link = f"http://localhost:8000/basic/test/intro/?test_id={test_uuid}"
    else:
        test_link = f"http://localhost:8000/basic/test/intro-sp/?test_id={test_uuid}"

    return JsonResponse({
        "message": "Test generated successfully.",
        "link": test_link,
        "copy_paste": f"Copy this link to access the test:\n{test_link}"
    })

def test_statistics(request, test_id):
    responses = Response.objects.filter(test_id=test_id)

    # Initialize variables to accumulate total latencies
    total_first_response_latency = 0
    total_all_latencies = 0
    total_latencies = 0  # Track the total number of latencies

    # Initialize counters for correct and incorrect responses
    correct_responses = 0
    incorrect_responses = 0

    # Loop through responses and their latencies
    for response in responses:
        latencies_string = response.latencies  # Assuming this is a string like "1129,570,506,459"

        if latencies_string:
            # Split the comma-separated string into a list of strings
            latencies = latencies_string.split(',')

            # Convert the latencies to integers
            latencies = [int(float(latency)) for latency in latencies] 

            # First latency (first response)
            first_response_latency = latencies[0]

            # Sum of all latencies (total response latency)
            total_response_latency = sum(latencies)

            # Accumulate the latencies
            total_first_response_latency += first_response_latency
            total_all_latencies += total_response_latency
            total_latencies += len(latencies)  # Add the number of latencies in this response

        # Count correct and incorrect responses
        if response.is_correct:
            correct_responses += 1
        else:
            incorrect_responses += 1

    # Calculate averages
    avg_first_response_latency = total_first_response_latency / len(responses) if len(responses) else 0
    avg_total_response_latency = total_all_latencies / total_latencies if total_latencies else 0

    # Prepare chart data for the bar chart
    chart_data = {
        "labels": ["Average First Response Latency", "Average Total Response Latency"],
        "values": [avg_first_response_latency, avg_total_response_latency],
    }

    # Prepare chart data for the pie chart
    pie_chart_data = {
        "labels": ["Correct Responses", "Incorrect Responses"],
        "values": [correct_responses, incorrect_responses],
    }

    return render(request, 'basic/dashboard/test_statistics.html', {
        "test_id": test_id,
        "chart_data": chart_data,
        "pie_chart_data": pie_chart_data,
    })






@csrf_exempt
def submit_all_responses(request):
    if request.method == "POST":
        data = json.loads(request.body)
        test_id = data["test_id"]

        test = TestSession.objects.get(test_id=test_id)
        test.state = "complete"
        test.date = timezone.now()
        test.save()

        correct_count = 0
        total_latency = 0
        latency_count = 0

        for item in data["responses"]:
            response = Response.objects.get(response_id=item["response_id"])
            response.response = item["response"]
            response.latencies = item["latencies"]
            try:
                latencies = [float(latency) for latency in item["latencies"].split(",")]
                latency = sum(latencies)
                amount = len(latencies)
                response.avg_latency = latency / amount
                total_latency += latency
                latency_count += 1
            except (ValueError, TypeError):
                pass  # skip if invalid

            # Determine correctness
            response.is_correct = (item["response"] == response.stim.correct_response)
            if response.is_correct:
                correct_count += 1

            response.save()

        # Now update the TestSession's performance stats
        total_responses = len(data["responses"])
        test.accuracy = correct_count / total_responses if total_responses > 0 else None
        test.avg_latency = total_latency / latency_count if latency_count > 0 else None
        test.save()

        return JsonResponse({"status": "success"})


from django.http import JsonResponse
from .models import Response


def get_responses(request):
    test_id = request.GET.get("test_id")
    responses = Response.objects.filter(test_id=test_id).select_related('stim')

    data = [
        {"response_id": r.response_id, "stimulus_text": r.stim.stimulus, "stimulus_type": r.stim.type}
        for r in responses
    ]

    return JsonResponse({"responses": data})

def logout_view(request):
    logout(request)
    return redirect("basic:login")

def test_intro(request):
    test_id = request.GET.get("test_id", None)
    return render(request, "basic/english_test/test_intro.html", {"test_id": test_id})

def test_instructions(request):
    test_id = request.GET.get("test_id", None)
    return render(request, "basic/english_test/test_instructions.html", {"test_id": test_id})

def practice_test(request):
    test_id = request.GET.get("test_id")

    return render(request, "basic/english_test/practice_test.html", {"test_id": test_id})

def get_practice_responses(request):
    practice_stimuli = Stimuli.objects.filter(type="Practice")[:2]
    data = [{"stimulus_text": s.stimulus} for s in practice_stimuli]
    return JsonResponse({"responses": data})

def practice_countdown(request):
    test_id = request.GET.get("test_id")
    return render(request, "basic/english_test/practice_countdown.html", {"test_id": test_id})

def practice_transition(request):
    test_id = request.GET.get("test_id")
    return render(request, "basic/english_test/practice_transition.html", {"test_id": test_id})

def take_test(request):
    test_id = request.GET.get("test_id")
    if not test_id:
        return HttpResponse("Error: Test ID missing", status=400)

    test = get_object_or_404(TestSession, pk=test_id)
    return render(request, "basic/english_test/test.html", {"test": test})

def test_complete(request):
    return render(request, "basic/english_test/test_complete.html")

def test_intro_sp(request):
    test_id = request.GET.get("test_id", None)
    return render(request, "basic/spanish_test/test_intro_sp.html", {"test_id": test_id})

def test_instructions_sp(request):
    test_id = request.GET.get("test_id", None)
    return render(request, "basic/spanish_test/test_instructions_sp.html", {"test_id": test_id})

def practice_test_sp(request):
    test_id = request.GET.get("test_id")

    return render(request, "basic/spanish_test/practice_test_sp.html", {"test_id": test_id})

def get_practice_responses_sp(request):
    practice_stimuli = Stimuli.objects.filter(type="Practice")[:2]
    data = [{"stimulus_text": s.stimulus} for s in practice_stimuli]
    return JsonResponse({"responses": data})

def practice_countdown_sp(request):
    test_id = request.GET.get("test_id")
    return render(request, "basic/spanish_test/practice_countdown_sp.html", {"test_id": test_id})

def practice_transition_sp(request):
    test_id = request.GET.get("test_id")
    return render(request, "basic/spanish_test/practice_transition_sp.html", {"test_id": test_id})

def take_test_sp(request):
    test_id = request.GET.get("test_id")
    if not test_id:
        return HttpResponse("Error: Test ID missing", status=400)

    test = get_object_or_404(TestSession, pk=test_id)
    return render(request, "basic/spanish_test/test_sp.html", {"test": test})

def test_complete_sp(request):
    return render(request, "basic/spanish_test/test_complete_sp.html")


def export_test_data(request):
    test_id = request.GET.get("test_id")
    test_session = get_object_or_404(TestSession, pk=test_id)
    responses = Response.objects.filter(test=test_session).select_related("stim")

    wb = Workbook()

    # Sheet 1: Raw Data
    raw_sheet = wb.active
    raw_sheet.title = "Raw Data"

    headers = ["Response ID", "Stimulus", "User Response", "Correct Response", "Is Correct", "Latencies (ms)"]
    raw_sheet.append(headers)

    correct_count = 0

    for resp in responses:
        is_correct_str = "Yes" if resp.is_correct else "No"
        raw_sheet.append([
            resp.response_id,
            resp.stim.stimulus,
            resp.response,
            resp.stim.correct_response,
            is_correct_str,
            resp.latencies,
        ])
        if resp.is_correct:
            correct_count += 1

    # Autofit columns
    for col in range(1, len(headers) + 1):
        raw_sheet.column_dimensions[get_column_letter(col)].width = 20

    # Sheet 2: Statistics
    stat_sheet = wb.create_sheet("Statistics")

    total_responses = responses.count()
    accuracy_percent = (correct_count / total_responses) * 100 if total_responses else 0
    avg_latency = test_session.avg_latency if test_session.avg_latency else 0

    stat_sheet.append(["Total Responses", total_responses])
    stat_sheet.append(["Correct Answers", correct_count])
    stat_sheet.append(["Accuracy (%)", round(accuracy_percent, 2)])
    stat_sheet.append(["Average Latency (ms)", avg_latency])
    stat_sheet.append([])
    stat_sheet.append(["Correctness Distribution", "Count"])
    stat_sheet.append(["Correct", correct_count])
    stat_sheet.append(["Incorrect", total_responses - correct_count])
    stat_sheet.append([])
    stat_sheet.append(["Stimulus", "Avg Latency (ms)"])

    # Average latency per stimulus
    latency_rows_start = stat_sheet.max_row + 1
    for r in responses:
        latency = r.avg_latency if r.avg_latency is not None else "N/A"
        stat_sheet.append([r.stim.stimulus, latency])

    # Charts

    # Pie Chart (Correct vs Incorrect)
    pie = PieChart()
    pie_labels = Reference(stat_sheet, min_col=1, min_row=7, max_row=8)
    pie_data = Reference(stat_sheet, min_col=2, min_row=7, max_row=8)
    pie.add_data(pie_data, titles_from_data=False)
    pie.set_categories(pie_labels)
    pie.title = "Correct vs Incorrect"
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showVal = True
    pie.dataLabels.showSerName = False
    stat_sheet.add_chart(pie, "E2")

    # Bar Chart (Latency per Stimulus)
    latency_chart = BarChart()
    latency_chart.title = "Average Latency per Stimulus"
    latency_chart.x_axis.title = "Stimulus"
    latency_chart.y_axis.title = "Latency (ms)"

    labels_start = latency_rows_start
    labels_end = stat_sheet.max_row
    labels = Reference(stat_sheet, min_col=1, min_row=labels_start, max_row=labels_end)
    data = Reference(stat_sheet, min_col=2, min_row=labels_start - 1, max_row=labels_end)

    latency_chart.add_data(data, titles_from_data=True)
    latency_chart.set_categories(labels)
    latency_chart.dataLabels = DataLabelList()
    latency_chart.dataLabels.showVal = True
    latency_chart.dataLabels.showCatName = False
    latency_chart.dataLabels.showSerName = False
    latency_chart.dataLabels.showLegendKey = False
    latency_chart.dataLabels.showPercent = False

    # Set chart size
    latency_chart.width = 20
    latency_chart.height = 12

    # Move legend to bottom
    latency_chart.legend.position = 'b'
    stat_sheet.add_chart(latency_chart, "E20")

    # Output the Excel file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="test_{test_id}_report.xlsx"'
    return response



from collections import defaultdict
from django.db.models import Count, Avg
from django.shortcuts import render
from .models import Response, TestSession

from collections import defaultdict
from django.db.models import Count, Avg
from django.shortcuts import render
from .models import Response, TestSession

def aggregated_statistics(request):
    responses = (
        Response.objects
        .exclude(latencies__isnull=True)
        .select_related("stim")
    )

    stim_sums = defaultdict(list)

    for r in responses:
        try:
            if r.avg_latency is not None:
                stim_sums[r.stim.stimulus].append(r.avg_latency)
        except (ValueError, AttributeError):
            continue

    stim_latency_data = [
        {"stim__stimulus": stim, "avg_latency": sum(vals) / len(vals)}
        for stim, vals in stim_sums.items()
    ]

    correctness_data = (
        Response.objects
        .exclude(is_correct__isnull=True)
        .values("is_correct")
        .annotate(count=Count("response_id"))
        .order_by("-is_correct")
    )

    # Group ages into 10-year buckets
    test_sessions = (
        TestSession.objects
        .exclude(age__isnull=True)
        .exclude(accuracy__isnull=True)
        .exclude(avg_latency__isnull=True)
        .values("age", "accuracy", "avg_latency")
    )


    age_brackets_accuracy = defaultdict(list)
    age_brackets_latency = defaultdict(list)

    for session in test_sessions:
        age = session["age"]
        bracket = (age // 10) * 10  # Grouping into 0-9, 10-19, etc.
        age_brackets_accuracy[bracket].append(session["accuracy"])
        age_brackets_latency[bracket].append(session["avg_latency"])

    age_labels = []
    age_accuracy = []
    age_latency = []

    for bracket in sorted(age_brackets_accuracy.keys()):
        accuracies = age_brackets_accuracy[bracket]
        latencies = age_brackets_latency[bracket]

        avg_accuracy = sum(accuracies) / len(accuracies) if accuracies else 0
        avg_latency = sum(latencies) / len(latencies) if latencies else 0

        age_labels.append(f"{bracket}-{bracket + 9}")
        age_accuracy.append(round(avg_accuracy * 100, 2))
        age_latency.append(round(avg_latency, 2))

    context = {
        "stim_labels": [entry["stim__stimulus"] for entry in stim_latency_data],
        "stim_latencies": [entry["avg_latency"] for entry in stim_latency_data],
        "correct_labels": ["Correct" if entry["is_correct"] else "Incorrect" for entry in correctness_data],
        "correct_counts": [entry["count"] for entry in correctness_data],
        "age_labels": age_labels,
        "age_accuracy": age_accuracy,
        "age_latency": age_latency,
    }

    return render(request, "basic/dashboard/aggregated_statistics.html", context)

